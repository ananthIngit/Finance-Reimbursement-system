from rest_framework import generics, permissions, status, views
from rest_framework.response import Response
from rest_framework_simplejwt.views import TokenObtainPairView
from django.contrib.auth import get_user_model
from django.db.models import Q, Count, Sum 
from rest_framework.exceptions import ValidationError 
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import filters
from django.db.models.functions import TruncMonth

# Import openpyxl and styling tools
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.http import HttpResponse
from django.utils.encoding import force_str
from django.utils.http import urlsafe_base64_decode
from rest_framework.pagination import PageNumberPagination

import random # Needed for OTP generation

# Import local modules
from .utils import send_status_update_email, send_otp_email 
from .models import Expense, ApprovalLog, Category, RolePermission, Permission, PasswordResetOTP, Role, Notification 
from .permissions import IsManager, IsFinance, IsSystemAdmin 
from .serializers import (
    UserRegistrationSerializer, ExpenseSerializer, CategorySerializer,
    CustomTokenObtainPairSerializer, UserProfileSerializer,
    ChangeUsernameSerializer, ChangeProfilePicSerializer,
    ChangePasswordSerializer, ExpenseActionSerializer,
    RoleSerializer, AdminUserListSerializer, AdminUserUpdateSerializer, NotificationSerializer 
)

User = get_user_model()

# ==========================
# 0. Custom Pagination Class
# ==========================
class CustomPagination(PageNumberPagination):
    page_size = 10 
    page_size_query_param = 'page_size' 
    max_page_size = 50 

# ==========================
# 1. Authentication Views
# ==========================

class CustomLoginView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer

class RegisterView(generics.CreateAPIView):
    serializer_class = UserRegistrationSerializer
    permission_classes = [permissions.AllowAny]

# ==========================
# 2. Expense Management
# ==========================

class CategoryListView(generics.ListAPIView):
    queryset = Category.objects.filter(is_active=True)
    serializer_class = CategorySerializer
    permission_classes = [permissions.IsAuthenticated]

class MyExpenseListView(generics.ListCreateAPIView):
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.IsAuthenticated]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    filterset_fields = {
        'status': ['exact'],
        'category': ['exact'],
        'created_at': ['gte', 'lte'],
    }
    
    search_fields = ['description', 'amount']
    ordering_fields = ['created_at', 'amount']

    def get_queryset(self):
        return Expense.objects.filter(employee=self.request.user)

    def perform_create(self, serializer):
        serializer.save(employee=self.request.user)

class MyExpenseDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Expense.objects.filter(employee=self.request.user)

    def perform_update(self, serializer):
        instance = self.get_object()
        if instance.status != 'Pending':
            raise ValidationError({"error": "You cannot edit an expense that has already been processed."})
        serializer.save()

    def perform_destroy(self, instance):
        if instance.status != 'Pending':
            raise ValidationError({"error": "You cannot delete an expense that has already been processed."})
        instance.delete()

# ==========================
# 3. Manager & Finance Logic
# ==========================

class TeamExpenseListView(generics.ListAPIView):
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.IsAuthenticated, IsManager]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    filterset_fields = {
        'status': ['exact'],
        'category': ['exact'],
        'created_at': ['gte', 'lte'],
    }
    
    search_fields = ['description', 'employee__username']
    ordering_fields = ['created_at', 'amount', 'employee__id']
    ordering = ['-created_at']

    def get_queryset(self):
        user = self.request.user
        return Expense.objects.filter(
            employee__department=user.department
        ).exclude(employee=user)

class FinanceQueueView(generics.ListAPIView):
    serializer_class = ExpenseSerializer
    permission_classes = [permissions.IsAuthenticated, IsFinance]
    pagination_class = CustomPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    filterset_fields = {
        'status': ['exact'],
        'category': ['exact'],
        'created_at': ['gte', 'lte'],
    }
    ordering_fields = ['created_at', 'amount', 'employee__id']
    ordering = ['-created_at']
    
    def get_queryset(self):
        return Expense.objects.all()

class ApproveRejectView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, pk):
        try:
            expense = Expense.objects.get(pk=pk)
        except Expense.DoesNotExist:
            return Response({"error": "Expense not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = ExpenseActionSerializer(data=request.data)
        if serializer.is_valid():
            action = serializer.validated_data['action']
            remarks = serializer.validated_data.get('remarks', '')
            user_role = request.user.role.name if request.user.role else ""

            if action in ['Approved', 'Rejected']:
                if expense.status != 'Pending':
                    return Response({"error": "Already processed"}, status=400)
                if user_role != 'Manager': 
                    return Response({"error": "Only Managers can approve"}, status=403)

            elif action == 'Reimbursed':
                if expense.status != 'Approved':
                    return Response({"error": "Must be approved first"}, status=400)
                if user_role != 'Finance':
                    return Response({"error": "Only Finance can reimburse"}, status=403)

            expense.status = action
            expense.save()
        
            ApprovalLog.objects.create(
                expense=expense, reviewer=request.user,
                new_status=action, remarks=remarks
            )
            
            # Create a notification for the employee
            Notification.objects.create(
                user=expense.employee,
                message=f"Your expense for {expense.category.name if expense.category else 'General'} (${expense.amount}) was {action.lower()} by {request.user.username}."
            )
            try:
                send_status_update_email(expense)
            except:
                pass
            return Response({"message": f"Expense marked as {action}", "status": action})
        return Response(serializer.errors, status=400)

# ==========================
# 4. Dashboards
# ==========================

class EmployeeDashboardView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]
    def get(self, request):
        queryset = Expense.objects.filter(employee=request.user)
        stats = queryset.aggregate(
            total_pending=Count('id', filter=Q(status='Pending')),
            total_approved=Count('id', filter=Q(status__in=['Approved', 'Reimbursed'])),
            total_rejected=Count('id', filter=Q(status='Rejected')),
            total_amount_claimed=Sum('amount', filter=Q(status='Reimbursed'))
        )
        stats['total_amount_claimed'] = stats['total_amount_claimed'] or 0
        return Response(stats)

class ManagerDashboardView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, IsManager]
    def get(self, request):
        team_expenses = Expense.objects.filter(
            employee__department=request.user.department
        ).exclude(employee=request.user)

        stats = team_expenses.aggregate(
            pending_approvals=Count('id', filter=Q(status='Pending')),
            total_approved=Count('id', filter=Q(status__in=['Approved', 'Reimbursed'])),
            total_rejected=Count('id', filter=Q(status='Rejected')),
            department_spend=Sum('amount', filter=Q(status='Reimbursed'))
        )
        stats['department_spend'] = stats['department_spend'] or 0
        return Response(stats)

class FinanceDashboardView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, IsFinance]
    def get(self, request):
        all_expenses = Expense.objects.all()
        stats = {
            "pending_payment": all_expenses.filter(status='Approved').count(),
            "total_paid": all_expenses.filter(status='Reimbursed').count(),
            "total_payout": all_expenses.filter(status='Reimbursed').aggregate(Sum('amount'))['amount__sum'] or 0
        }
        return Response(stats)

# ==========================
# 5. Profile & Export
# ==========================

class MyProfileView(generics.RetrieveUpdateAPIView):
    """View to see and update profile including Bank Details"""
    serializer_class = UserProfileSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser, JSONParser] 
    def get_object(self): return self.request.user

class ChangeUsernameView(generics.UpdateAPIView):
    serializer_class = ChangeUsernameSerializer
    permission_classes = [permissions.IsAuthenticated]
    def get_object(self): return self.request.user

class ChangeProfilePicView(generics.UpdateAPIView):
    serializer_class = ChangeProfilePicSerializer
    permission_classes = [permissions.IsAuthenticated]
    def get_object(self): return self.request.user

class ChangePasswordView(generics.UpdateAPIView):
    serializer_class = ChangePasswordSerializer
    permission_classes = [permissions.IsAuthenticated]
    def get_object(self): return self.request.user
    def update(self, request, *args, **kwargs):
        user = self.get_object()
        serializer = self.get_serializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            user.set_password(serializer.validated_data['new_password'])
            user.save()
            return Response({"message": "Password updated"}, status=200)
        return Response(serializer.errors, status=400)

class ExportExpensesView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, IsFinance]
    
    def get(self, request):
        status_filter = request.query_params.get('status')
        queryset = Expense.objects.all().order_by('-created_at')
        if status_filter: 
            queryset = queryset.filter(status=status_filter)
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expense Report"

        headers = ["Sl No.", "Employee", "Category", "Amount", "Status", "Date"]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        center_aligned = Alignment(horizontal="center", vertical="center")
        left_aligned = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'), 
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        for col_num, cell in enumerate(ws[1], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_aligned
            cell.border = thin_border

        column_widths = {'A': 10, 'B': 25, 'C': 25, 'D': 15, 'E': 18, 'F': 20}
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width

        for index, exp in enumerate(queryset, start=1):
            date_str = exp.created_at.strftime("%Y-%m-%d")
            row_data = [index, exp.employee.username, exp.category.name, float(exp.amount), exp.status, date_str]
            ws.append(row_data)
            
            current_row = ws.max_row
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
                cell = ws[f"{col_letter}{current_row}"]
                cell.border = thin_border
                cell.alignment = center_aligned if col_letter in ['A', 'D', 'E', 'F'] else left_aligned
                if col_letter == 'D':
                    cell.number_format = '"$"#,##0.00'

        ws.freeze_panes = 'A2'
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        filename_date = queryset.first().created_at.strftime("%Y%m%d") if queryset.exists() else "Empty"
        response['Content-Disposition'] = f'attachment; filename="Expense_Report_{filename_date}.xlsx"'
        wb.save(response)
        return response

class DeleteUserView(generics.DestroyAPIView):
    queryset = User.objects.all()
    permission_classes = [permissions.IsAuthenticated]
    def delete(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance == request.user: return Response({"error": "Cannot delete self"}, status=400)
        instance.is_active = False 
        instance.save()
        return Response({"message": "User deactivated"}, status=200)

# ==========================
# 6. OTP Password Reset Views
# ==========================

class RequestOTPView(views.APIView):
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        email = request.data.get('email')
        if not email:
            return Response({'error': 'Email is required.'}, status=400)

        user = User.objects.filter(email=email).first()
        
        if user:
            otp_code = str(random.randint(100000, 999999))
            PasswordResetOTP.objects.create(user=user, otp=otp_code)
            
            # Sending real email now that App Password is set up
            send_otp_email(user, otp_code)
            
        return Response({'message': 'If an account with that email exists, an OTP has been sent.'}, status=200)

class ResetPasswordWithOTPView(views.APIView):
    permission_classes = [permissions.AllowAny]
    
    def post(self, request):
        email = request.data.get('email')
        otp_code = request.data.get('otp')
        new_password = request.data.get('new_password')

        if not all([email, otp_code, new_password]):
            return Response({'error': 'Email, OTP, and new password are required.'}, status=400)

        if len(new_password) < 6:
            return Response({'error': 'Password must be at least 6 characters long.'}, status=400)

        user = User.objects.filter(email=email).first()
        if not user:
            return Response({'error': 'Invalid request.'}, status=400)

        otp_record = PasswordResetOTP.objects.filter(user=user, otp=otp_code).order_by('-created_at').first()

        if not otp_record or not otp_record.is_valid():
            return Response({'error': 'Invalid or expired OTP code.'}, status=400)

        user.set_password(new_password)
        user.save()
        PasswordResetOTP.objects.filter(user=user).delete()

        return Response({'message': 'Password has been successfully reset!'}, status=200)
    
# Add this class to your views.py
class VerifyOTPView(views.APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get('email')
        otp_code = request.data.get('otp')

        user = User.objects.filter(email=email).first()
        if not user:
            return Response({'error': 'Invalid request.'}, status=400)

        otp_record = PasswordResetOTP.objects.filter(user=user, otp=otp_code).order_by('-created_at').first()

        if not otp_record or not otp_record.is_valid():
            return Response({'error': 'Invalid or expired OTP code.'}, status=400)

        return Response({'message': 'OTP verified successfully!'}, status=200)

# ==========================
# 7. Admin Management Views
# ==========================

class AdminUserListView(generics.ListAPIView):
    queryset = User.objects.all().order_by('-date_joined')
    serializer_class = AdminUserListSerializer
    permission_classes = [permissions.IsAuthenticated, IsSystemAdmin]
    pagination_class = CustomPagination
    filter_backends = [filters.SearchFilter, DjangoFilterBackend]
    search_fields = ['username', 'email', 'department']
    filterset_fields = ['is_active', 'role']

class AdminUserDetailView(generics.RetrieveUpdateAPIView):
    queryset = User.objects.all()
    serializer_class = AdminUserUpdateSerializer
    permission_classes = [permissions.IsAuthenticated, IsSystemAdmin]

class AdminDropdownDataView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, IsSystemAdmin]
    
    def get(self, request):
        roles = Role.objects.values('id', 'name')
        managers = User.objects.filter(role__name__in=['Manager', 'Admin'], is_active=True).values('id', 'username', 'email')
        return Response({'roles': list(roles), 'managers': list(managers)})

class AdminStatsView(views.APIView):
    permission_classes = [permissions.IsAuthenticated, IsSystemAdmin]
    def get(self, request):
        users = User.objects.all()
        return Response({
            "total_users": users.count(),
            "active_users": users.filter(is_active=True).count(),
            "unassigned_roles": users.filter(role__isnull=True).count(),
            "deactivated": users.filter(is_active=False).count()
        })

# ==========================
# 8. Notifications
# ==========================

class NotificationListView(generics.ListAPIView):
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return Notification.objects.filter(user=self.request.user).order_by('-created_at')

class MarkNotificationReadView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request, pk):
        try:
            notification = Notification.objects.get(pk=pk, user=request.user)
            notification.is_read = True
            notification.save()
            return Response({"message": "Notification marked as read"})
        except Notification.DoesNotExist:
            return Response({"error": "Notification not found"}, status=404)

class MarkAllNotificationsReadView(views.APIView):
    permission_classes = [permissions.IsAuthenticated]

    def patch(self, request):
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({"message": "All notifications marked as read"})